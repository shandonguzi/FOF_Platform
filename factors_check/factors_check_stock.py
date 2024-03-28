
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
from settings.database import *
from utils.mysql.get_sql import get_sql
from utils.time_function import timeit
import warnings
warnings.filterwarnings('ignore')


# Function to remove outliers using IQR
def remove_outliers_with_iqr(df, column):
    Q1 = df[column].quantile(0.25)
    Q3 = df[column].quantile(0.75)
    IQR = Q3 - Q1
    lower_bound = Q1 - 1.5 * IQR
    upper_bound = Q3 + 1.5 * IQR
    return df[(df[column] >= lower_bound) & (df[column] <= upper_bound)]


def calculate_returns(group, n_quantiles):
    # 计算长仓和短仓股票的平均收益率
    long_stocks = group[group['quantile'] == 0]['Stkcd']
    short_stocks = group[group['quantile'] == (n_quantiles - 1)]['Stkcd']
    long_return = group[group['Stkcd'].isin(long_stocks)]['RealPctChange'].mean()
    short_return = group[group['Stkcd'].isin(short_stocks)]['RealPctChange'].mean()*(-1)
    long_short_return = long_return + short_return
    
    return pd.Series({
        'Long_Return': long_return,
        'Short_Return': short_return,
        'Long_Short_Return': long_short_return
    })

def annualized_return(returns_series):
    compounded_growth = (1 + returns_series).prod()
    n_periods = len(returns_series)
    return compounded_growth**(12/n_periods) - 1

# 定义计算夏普比率的函数
def sharpe_ratio(returns_series, risk_free_rate=0):
    # 默认无风险利率为0，实际使用时应根据实际情况调整
    excess_returns = returns_series - risk_free_rate/12
    return excess_returns.mean() / excess_returns.std() * np.sqrt(12)

# 定义计算最大回撤的函数
def max_drawdown(returns_series):
    cumulative_returns = (1 + returns_series).cumprod()
    peak = cumulative_returns.expanding(min_periods=1).max()
    drawdown = (cumulative_returns/peak) - 1
    return drawdown.min()

# Backtest function to evaluate factors
def backtest_factors(factor_data, month_return, zz_500, n_quantiles, time_start, time_end=None, direction='high'):
    
    # Merge and filter the data
    merged_data = pd.merge(factor_data, month_return, on=['Date', 'Stkcd'], how='inner')
    merged_data = merged_data[(merged_data.Date >= time_start) & (merged_data.Date <= time_end)] if time_end else merged_data[merged_data.Date >= time_start]

    # Sort stocks based on factor value and directionality
    if direction == 'high':  # High values are good
        merged_data['quantile'] = pd.qcut(merged_data[factor_data.columns[2]], n_quantiles, labels=False)
    else:  # Low values are good
        merged_data['quantile'] = pd.qcut(-merged_data[factor_data.columns[2]], n_quantiles, labels=False)

    # Calculate returns for quantile groups
    returns = merged_data.groupby('Date').apply(calculate_returns, n_quantiles=n_quantiles)

    # Performance metrics
    eval_metrics = {
        'Annualized Return': annualized_return(returns['Long_Short_Return']),
        'Sharpe Ratio': sharpe_ratio(returns['Long_Short_Return']),
        'Max Drawdown': max_drawdown(returns['Long_Short_Return'])
    }

    # Plotting returns against benchmark
    returns['L_Cumulative'] = (1 + returns['Long_Return']).cumprod() - 1
    returns['S_Cumulative'] = (1 + returns['Short_Return']).cumprod() - 1
    returns['LS_Cumulative'] = (1 + returns['Long_Short_Return']).cumprod() - 1

    zz500_copy = zz_500.copy()
    zz500_copy.index = zz500_copy.index - pd.offsets.MonthBegin(1)
    zz500_copy = zz500_copy[zz500_copy.index.isin(returns.index)]
    zz500_copy.loc[:, 'Cumulative'] = (1 + zz500_copy['ed2ed']).cumprod() - 1

    # Plot code here...
    plot_filename = f"/code/factors_check/{factor_data.columns[2]}_cumulative_returns.png"
    plt.figure(figsize=(10, 5))
    plt.plot(returns.index, returns['LS_Cumulative'], label='Long-Short Portfolio')
    plt.plot(returns.index, returns['L_Cumulative'], label='Long Portfolio')
    plt.plot(returns.index, returns['S_Cumulative'], label='Short Portfolio')
    plt.plot(returns.index, zz500_copy['Cumulative'], label='ZZ500 Index')
    plt.legend()
    plt.title(f"Cumulative Returns compared to ZZ500 - {factor_data.columns[2]}")
    plt.xlabel('Date')
    plt.ylabel('Cumulative Returns')
    plt.tight_layout()
    plt.savefig(plot_filename)
    plt.close()

    return returns, eval_metrics, plot_filename

def output_to_excel(all_factor_results, output_file_path):
    # Create a workbook and add sheets
    wb = Workbook()
    # Remove default sheet if it exists
    if 'Sheet' in wb.sheetnames:
        std = wb['Sheet']
        wb.remove(std)

    for factor, (returns, eval_metrics, plot_filename) in all_factor_results.items():
        # Each factor's analysis results go into its own sheet
        ws = wb.create_sheet(title=f'{factor}_Analysis')

        # Insert returns data into the sheet
        for r in dataframe_to_rows(returns, index=True, header=True):
            ws.append(r)
        
        # Add evaluation metrics below the returns data with some space
        metrics_start_row = ws.max_row + 3
        for metric_name, value in eval_metrics.items():
            ws.cell(row=metrics_start_row, column=1, value=metric_name).font = Font(bold=True)
            ws.cell(row=metrics_start_row, column=2, value=value)
            metrics_start_row += 1
        
        # Load and add the plot image to the sheet
        img = OpenpyxlImage(plot_filename)
        # You might need to adjust this depending on where you want the image
        ws.add_image(img, 'A' + str(metrics_start_row + 2))

    # Save the workbook
    wb.save(output_file_path)


@timeit('Factors check')
def factors_check():
    Stock_Month_Return = get_sql(level1_csmar, 'TRD_Mnth')
    Stock_Month_Return = Stock_Month_Return[['Date', 'Stkcd', 'RealPctChange']]
    zz_500 = get_sql(level1_csmar, 'zz500_monthly')
    zz_500 = zz_500[['Date', 'ed2ed']].set_index('Date')

    factors = [('IdioVol', 'high'),  # Trading Friction
               ('Rel2High', 'high'),
               ('ResidVar', 'high'),
               ('AT', 'low'),
               ('CAPM_beta', 'high'),
               ('LME', 'low'),
               ('LTurnover', 'high'),
               ('MktBeta', 'high'),
            #    ('Variance', 'high')]
               ('r2_1', 'high'),  # past return
               ('r12_2', 'high'),
               ('r12_7', 'high'),
               ('r36_13', 'high'),
               ('Investment', 'high'),  # Investment
               ('NOA', 'high'),
               ('DPI2A', 'high'),
               ('NSI', 'high'),
               ('PROF', 'high'),  # Profitability
               ('ATO', 'high'),
               ('CTO', 'high'),
               ('FC2Y', 'high'),
               ('OP', 'high'),
               ('PM', 'high'),
               ('RNA', 'high'),
               ('ROA', 'high'),
               ('ROE', 'high'),
               ('SGA2S', 'high'),
               ('D2A', 'high'),
               ('AC', 'high'),  # Intangibles
               ('OL', 'high'),
               ('PCM', 'high'),
               ('A2ME', 'high'),  # Value
               ('BEME', 'high'),
               ('C', 'high'),
               ('CF', 'high'),
               ('CF2P', 'high'),
               ('E2P', 'high'),
               ('Q', 'high'),
               ('S2P', 'high'),
               ('Lev', 'high'),
    ]
    analysis_results = {}

    for factor, direction in factors:
        factor_data = get_sql(level3_factors, factor)
        factor_data['Date'] = pd.to_datetime(factor_data['Date']) + pd.offsets.MonthBegin(1)
        factor_data = remove_outliers_with_iqr(factor_data, factor_data.columns[2])
        portfolio_returns, eval_metrics, plot_filename = backtest_factors(factor_data, Stock_Month_Return, zz_500, 5, '2013-01-01', direction=direction)
        analysis_results[factor] = (portfolio_returns, eval_metrics, plot_filename)
        print(f'Finished analyzing factor: {factor}')

    # Call output to Excel function to save all results in one file
    output_to_excel(analysis_results, '/code/factors_check/all_factors_analysis.xlsx')